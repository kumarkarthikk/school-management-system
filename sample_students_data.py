# sample_students_data.py

from openpyxl import load_workbook
import random
from datetime import datetime, timedelta

# ==========================================
# LOAD EXCEL FILE
# ==========================================

EXCEL_FILE = "students.xlsx"

wb = load_workbook(EXCEL_FILE)
sheet = wb.active

# ==========================================
# LOCATIONS
# ==========================================

locations = {
    "Visakhapatnam": [
        "Gajuwaka",
        "Madhurawada",
        "Anakapalle",
        "Bheemunipatnam",
        "Pendurthi",
        "NAD",
        "Dwaraka Nagar",
        "Seethammadhara"
    ],

    "Vizianagaram": [
        "Cheepurupalli",
        "Bobbili",
        "Nellimarla",
        "Salur",
        "Gajapathinagaram",
        "Parvathipuram",
        "Kothavalasa"
    ],

    "Srikakulam": [
        "Palasa",
        "Tekkali",
        "Ichapuram",
        "Narasannapeta",
        "Rajam",
        "Amadalavalasa",
        "Pathapatnam"
    ]
}

# ==========================================
# STUDENT FIRST NAMES
# ==========================================

first_names = [
    "Sai", "Karthik", "Rohit", "Vamsi", "Nikhil",
    "Harsha", "Dinesh", "Ajay", "Teja", "Ganesh",
    "Rakesh", "Suresh", "Charan", "Pavan", "Yash",
    "Abhi", "Mahesh", "Naresh", "Lokesh", "Praveen",
    "Arjun", "Tarun", "Rahul", "Surya", "Keerthana",
    "Anusha", "Sravani", "Bhavya", "Divya", "Navya",
    "Harini", "Deepika", "Sneha", "Tejaswini",
    "Sushmitha", "Mounika", "Lahari", "Poojitha"
]

# ==========================================
# FAMILY SURNAME LIST
# ==========================================

family_last_names = [
    "Kumar",
    "Reddy",
    "Naidu",
    "Rao",
    "Varma",
    "Yadav",
    "Patro",
    "Panda",
    "Nayak",
    "Murthy"
]

# ==========================================
# CREATE FAMILIES
# ==========================================

families = []

for i in range(70):

    surname = random.choice(family_last_names)

    father_first = random.choice([
        "Ramesh",
        "Srinivas",
        "Prasad",
        "Murali",
        "Krishna",
        "Ganesh",
        "Mohan",
        "Suresh",
        "Ravi",
        "Naresh",
        "Rajesh",
        "Harish",
        "Venkatesh"
    ])

    mother_first = random.choice([
        "Lakshmi",
        "Padma",
        "Anitha",
        "Sujatha",
        "Sravani",
        "Bhavani",
        "Kavitha",
        "Jyothi",
        "Sunitha",
        "Deepa"
    ])

    father_name = father_first + " " + surname
    mother_name = mother_first + " " + surname

    phone = "9" + str(random.randint(100000000, 999999999))

    families.append({
        "surname": surname,
        "father_name": father_name,
        "mother_name": mother_name,
        "phone": phone
    })

# ==========================================
# PREVIOUS SCHOOLS
# ==========================================

schools = [
    "Sri Chaitanya School",
    "Narayana School",
    "ZP High School",
    "Government School",
    "Oxford School",
    "Little Flowers School",
    "Vignan School",
    "Delhi Public School",
    "St Anns School"
]

# ==========================================
# REMARKS
# ==========================================

remarks_list = [
    "Good Student",
    "Excellent in Maths",
    "Sports Player",
    "Needs Improvement",
    "Very Active",
    "Regular Attendance",
    "Top Performer",
    "Silent Student",
    "Discipline Good"
]

# ==========================================
# CLASS LIST
# ==========================================

classes = [
    "1st",
    "2nd",
    "3rd",
    "4th",
    "5th",
    "6th",
    "7th",
    "8th",
    "9th",
    "10th"
]

# ==========================================
# NUMBER OF STUDENTS
# ==========================================

student_count = 187

# ==========================================
# START INSERTING
# ==========================================

start_row = sheet.max_row + 1

for i in range(student_count):

    row = start_row + i

    student_id = row - 1

    # RANDOM DATE
    random_days = random.randint(0, 500)

    admission_date = (
        datetime.today() - timedelta(days=random_days)
    ).strftime("%Y-%m-%d")

    # SELECT FAMILY
    family = random.choice(families)

    father_name = family["father_name"]
    mother_name = family["mother_name"]
    phone = family["phone"]
    surname = family["surname"]

    # STUDENT NAME
    first_name = random.choice(first_names)

    student_name = first_name + " " + surname

    # CLASS
    student_class = random.choice(classes)

    # CITY + TOWN
    city = random.choice(list(locations.keys()))

    town = random.choice(locations[city])

    full_location = town + ", " + city

    # SCHOOL
    previous_school = random.choice(schools)

    # STATUS
    status = "ACTIVE"

    # REMARKS
    remarks = random.choice(remarks_list)

    # ==========================================
    # INSERT INTO EXCEL
    # ==========================================

    sheet.cell(row, 1).value = student_id
    sheet.cell(row, 2).value = admission_date
    sheet.cell(row, 3).value = student_name
    sheet.cell(row, 4).value = student_class
    sheet.cell(row, 5).value = father_name
    sheet.cell(row, 6).value = mother_name
    sheet.cell(row, 7).value = full_location
    sheet.cell(row, 8).value = phone
    sheet.cell(row, 9).value = previous_school
    sheet.cell(row, 10).value = status
    sheet.cell(row, 11).value = remarks

# ==========================================
# SAVE FILE
# ==========================================

wb.save(EXCEL_FILE)

print("\n===================================")
print("187 STUDENT RECORDS ADDED SUCCESSFULLY")
print("students.xlsx UPDATED SUCCESSFULLY")
print("===================================")