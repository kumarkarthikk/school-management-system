from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
import os, datetime

app = Flask(__name__, static_folder='static')
CORS(app)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'students.xlsx')

COLS = {
    'id': 1, 'date': 2, 'name': 3, 'class': 4,
    'father_name': 5, 'mother_name': 6, 'city': 7,
    'phone': 8, 'previous_school': 9, 'status': 10, 'remarks': 11
}

def row_to_dict(sheet, row):
    return {
        'id':              sheet.cell(row, COLS['id']).value,
        'date':            str(sheet.cell(row, COLS['date']).value or ''),
        'name':            sheet.cell(row, COLS['name']).value,
        'class':           sheet.cell(row, COLS['class']).value,
        'father_name':     sheet.cell(row, COLS['father_name']).value,
        'mother_name':     sheet.cell(row, COLS['mother_name']).value,
        'city':            sheet.cell(row, COLS['city']).value,
        'phone':           str(sheet.cell(row, COLS['phone']).value or ''),
        'previous_school': sheet.cell(row, COLS['previous_school']).value,
        'status':          sheet.cell(row, COLS['status']).value,
        'remarks':         sheet.cell(row, COLS['remarks']).value,
    }

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/students', methods=['GET'])
def get_students():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    students = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, COLS['id']).value is not None:
            students.append(row_to_dict(sheet, row))
    return jsonify(students)

@app.route('/api/students/search', methods=['GET'])
def search_students():
    field = request.args.get('field', 'name')
    value = request.args.get('value', '').strip().lower()
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    results = []
    col_map = {'name': COLS['name'], 'father_name': COLS['father_name'], 'phone': COLS['phone']}
    col = col_map.get(field, COLS['name'])
    for row in range(2, sheet.max_row + 1):
        cell_val = str(sheet.cell(row, col).value or '').strip().lower()
        if value in cell_val:
            results.append(row_to_dict(sheet, row))
    return jsonify(results)

@app.route('/api/students', methods=['POST'])
def add_student():
    data = request.json
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    row = sheet.max_row + 1
    student_id = row - 1
    sheet.cell(row, COLS['id']).value = student_id
    sheet.cell(row, COLS['date']).value = data.get('date', str(datetime.date.today()))
    sheet.cell(row, COLS['name']).value = data.get('name')
    sheet.cell(row, COLS['class']).value = data.get('class')
    sheet.cell(row, COLS['father_name']).value = data.get('father_name')
    sheet.cell(row, COLS['mother_name']).value = data.get('mother_name')
    sheet.cell(row, COLS['city']).value = data.get('city')
    sheet.cell(row, COLS['phone']).value = data.get('phone')
    sheet.cell(row, COLS['previous_school']).value = data.get('previous_school')
    sheet.cell(row, COLS['status']).value = 'ACTIVE'
    sheet.cell(row, COLS['remarks']).value = data.get('remarks', '')
    wb.save(EXCEL_PATH)
    return jsonify({'success': True, 'id': student_id})

@app.route('/api/students/<int:student_id>', methods=['PUT'])
def update_student(student_id):
    data = request.json
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, COLS['id']).value == student_id:
            for field, col in COLS.items():
                if field in data and field != 'id':
                    sheet.cell(row, col).value = data[field]
            wb.save(EXCEL_PATH)
            return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Student not found'}), 404

@app.route('/api/students/<int:student_id>/leave', methods=['POST'])
def leave_student(student_id):
    data = request.json
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, COLS['id']).value == student_id:
            if str(sheet.cell(row, COLS['status']).value).upper() == 'LEFT':
                return jsonify({'success': False, 'error': 'Student already marked as LEFT'}), 400
            sheet.cell(row, COLS['status']).value = 'LEFT'
            reason = data.get('reason', '')
            remarks = data.get('remarks', '')
            sheet.cell(row, COLS['remarks']).value = f"LEFT STUDENT | Reason: {reason} | Remarks: {remarks}"
            wb.save(EXCEL_PATH)
            return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Student not found'}), 404

@app.route('/api/stats', methods=['GET'])
def get_stats():
    wb = load_workbook(EXCEL_PATH)
    sheet = wb.active
    total = active = left = 0
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, COLS['id']).value is not None:
            total += 1
            status = str(sheet.cell(row, COLS['status']).value or '').upper()
            if status == 'ACTIVE': active += 1
            elif status == 'LEFT': left += 1
    return jsonify({'total': total, 'active': active, 'left': left})

if __name__ == '__main__':
    os.makedirs('static', exist_ok=True)
    app.run(debug=True, port=5000)